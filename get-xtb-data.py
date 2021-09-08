import pandas_datareader.data as web
import pandas as pd
from datetime import datetime, timedelta
import numpy as np
from openpyxl import load_workbook
import os
import dropbox
from dropbox.files import WriteMode


#  Defining analytical functions
def price_change(current_price, old_price):
    
    price_change = (current_price - old_price)/current_price
    return price_change

## Implementing search for constant price decrease in period n
## The idea is to use sorting in a descending order 
## If sorted list is identical to the one coming from the n period - it means the price has constantly been dropping

def constant_price_drop_detector(df_ticker, n_period):
    prices_actual = df_ticker.tail(n_period)['Adj Close']
    prices_sorted = prices_actual.sort_values(ascending = False)

    ## Now we check whether the lists are identical. The operation returns an array of True, Flase values. 
    ## Using np.mean we can verify whether all the list elements are True -> Then mean = 1

    array_mean = np.mean(prices_actual.values == prices_sorted.values)

    if array_mean == 1:
        constant_price_drop = "YES"
    else:
        constant_price_drop = "NO"
    
    return constant_price_drop


def get_data(ticker, tickers):
    """
    Function used to get data for a particular ticker from yahoo - using pandas webreader
    """
    
    print(ticker)
    ## Date setting
    today = datetime.today()
    days_ago_90 = today  - timedelta(days = 90)
    today = today.strftime("%Y-%m-%d")
    days_ago_90 = days_ago_90.strftime("%Y-%m-%d")
    
    df_ticker = web.DataReader(ticker, 'yahoo', start = days_ago_90, end = today)
    
    ## To get prices, iloc is used. It's because shifting by timedetlas will result in error in cases where some holidays occured    
    price_most_recent = df_ticker.iloc[-1, 5]
    price_7_days_ago = df_ticker.iloc[-7, 5]
    price_21_days_ago = df_ticker.iloc[-21, 5]
    price_30_days_ago = df_ticker.iloc[-30, 5]
    price_90_days_ago = df_ticker.iloc[0,5]
    
    ## Getting price change
    price_change_7_days = price_change(price_most_recent, price_7_days_ago)
    price_change_21_days = price_change(price_most_recent, price_21_days_ago)
    price_change_30_days = price_change(price_most_recent, price_30_days_ago)
    price_change_90_days = price_change(price_most_recent, price_90_days_ago)
    
    ## Checking for constant price drop
    constant_price_drop_7 = constant_price_drop_detector(df_ticker, 7)
    ## Only if price drops constantly for 7 days it makes sense to check for this pattern in 21 days period
    if constant_price_drop_7 == "YES":
        constant_price_drop_21 = constant_price_drop_detector(df_ticker, 21)
    else:
        constant_price_drop_21 = "NO"
    
    ## Now creating the final df to return
    df_prices = df_ticker[['Adj Close']].T
    df_prices.index = [ticker]
    df_prices.reset_index(inplace = True)
    
    full_name = tickers.loc[tickers["Ticker"] == ticker, 'Full Name'].values[0]
    df_prices['company_name'] = full_name
    df_prices['price_90_days_ago'] = price_90_days_ago
    df_prices['price_30_days_ago'] = price_30_days_ago
    df_prices['price_21_days_ago'] = price_21_days_ago
    df_prices['price_7_days_ago'] = price_7_days_ago
    df_prices['price_most_recent'] = price_most_recent
    
    df_prices['price_change_7_days'] = price_change_7_days
    df_prices['price_change_21_days'] = price_change_21_days
    df_prices['price_change_30_days'] = price_change_30_days
    df_prices['price_change_90_days'] = price_change_90_days
    
    df_prices['constant_price_drop_7'] = constant_price_drop_7
    df_prices['constant_price_drop_21'] = constant_price_drop_21
    
    df_prices.fillna("None", inplace = True)
    
    return df_prices


## Formatting the spreadsheet

def format_sheet(sheet_to_be_formatted, output_name):

    """Function used to set a desired format to already created excel file with stock data"""

    wb = load_workbook(sheet_to_be_formatted)
    # ws = book['Sheet1']
    ws = wb.active

    columns_to_keep_unhidden = ['index', 'company_name','price_90_days_ago', 'price_30_days_ago', 'price_21_days_ago', 
                                'price_7_days_ago', 'price_most_recent', 
                               'price_change_7_days', 'price_change_21_days', 'price_change_30_days', 'price_change_90_days', 
                               'constant_price_drop_7', 'constant_price_drop_21']

    ## Columns to be formated
    percentage_columns = ['price_change_7_days', 'price_change_21_days', 'price_change_30_days', 'price_change_90_days']

    ## Creating dicts
    percentage_columns_dict = {}
    
    columns_to_keep_dict = {}
    for column_cell in ws.iter_cols(1, ws.max_column):  # iterate column cell

        ## Hiding daily prices columns

        if column_cell[0].value not in columns_to_keep_unhidden:
            col = column_cell[0].column_letter
            ws.column_dimensions[col].hidden= True

        if column_cell[0].value in percentage_columns:
            percentage_columns_dict[column_cell[0].value] = column_cell[0].column_letter
            
        if column_cell[0].value in columns_to_keep_unhidden:
            columns_to_keep_dict[column_cell[0].value] = column_cell[0].column_letter
    
    
    # Percentage format can be only set to cells. Thus we need to iterrate through rows of applicable columns
    row_nums = len(list(ws.rows))
    print(row_nums)
    for v in percentage_columns_dict.values():
        for row_num in range(2, row_nums):
            ws[f'{v}{row_num}'].number_format = '0.00%'
    
#     Setting width for selected columns
    for k in columns_to_keep_dict.keys():
        col = columns_to_keep_dict[k]
        print(col)
        ws.column_dimensions[col].width = 20



    wb.save(output_name)


def upload_dropbox():
    """Function used to upload processed data to dropbox for easy access - as scrip will be run daily on raspberry pi"""

    # Hide access token! 
    dropbox_access_token = os.environ['DROPBOX']
    dropbox_path = '/xtb-day-trading/stocks.xlsx'

    # Computer path to be adjusted
    curr_dir = os.getcwd()
    computer_path = curr_dir + "\\stocks_formatted.xlsx"

    client = dropbox.Dropbox(dropbox_access_token)
    print("[SUCCESS] dropbox account linked")

    client.files_upload(open(computer_path, "rb").read(), dropbox_path, mode=WriteMode('overwrite'))
    print("[UPLOADED] {}".format(computer_path))


def run():

    # Getting tickers from a pre-defined file
    tickers = pd.read_csv("xtb_tickers.csv", header = None)
    tickers.columns = ["Ticker", "Full Name"]
    tickers['Ticker'] = tickers['Ticker'].apply(lambda x: x.replace(".US",""))

    # Getting actual data
    ## Method without openpyxl and writing next rows to excel - this doesn't seem to work properly
    for ticker in tickers['Ticker']: 
        try:
            df_prices = get_data(ticker, tickers)
            df = df.append(df_prices)
        except:
            try:
                df_prices = get_data(ticker, tickers)
                df = df_prices.copy()
            except:
                print("Error:", ticker)
                continue


    df.to_excel("stocks.xlsx", index = False)

    # Formatting excel file produced in a previous step
    format_sheet('stocks.xlsx', 'stocks_formatted.xlsx')

    # Uploading to dropbox
    upload_dropbox()

run()

